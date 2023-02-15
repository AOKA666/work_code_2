import openpyxl


def get_teacher_dict():
    wb = openpyxl.load_workbook("老师信息.xlsx")
    ws = wb.worksheets[0]
    nrow = ws.max_row
    
    result = {}
    for i in range(2, nrow+1):
        name = ws.cell(i,1).value
        t_id = str(ws.cell(i,2).value)
        result[name] = t_id
    return result


if __name__ == '__main__':
    t = get_teacher_dict()
    print(list(t.values()))
    new_dict = dict(zip(t.values(), t.keys()))
    print(new_dict)