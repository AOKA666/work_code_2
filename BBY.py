import openpyxl
from crm import FM
from excel import Excel
import os
import warnings
warnings.filterwarnings('ignore')


def download_excel(download_path):
    """获取下载的石墨表格"""
    origin_path = os.getcwd()
    os.chdir(download_path)
    item = os.listdir()
    target_file = []
    for i in item:
        filename = '{}\{}'.format(download_path, i)
        if os.path.isfile(filename):
            if '实验班' in i:
                target_file.append(filename)
    if not target_file:
        return
    newest = 0
    document = ''
    for f in target_file:
        modify_time = os.path.getmtime(f)
        if modify_time > newest:
            newest = modify_time
            document = f
    os.chdir(origin_path)
    return document

def get_data(crm_user, crm_pwd):
    fm = FM(crm_user, crm_pwd)
    return fm.get_list()

def search_in_excel(document_name, crm_user, crm_pwd):
    """获取学员列表，根据学员列表处理石墨表格"""
    data = get_data(crm_user, crm_pwd)
    excel = Excel(document_name)
    print("\n开始更新excel表格...")
    allocated = excel.search(data)
    return allocated

def deploy(allocated, crm_user, crm_pwd):
    stu_list = get_data(crm_user, crm_pwd)
    fm = FM(crm_user, crm_pwd)
    diploy_list = fm.cal_number(stu_list, allocated)
    s = input("\n检查一下有没有问题，没问题就分配啦(按y)>>>")
    if s.lower() == 'y':
        fm.assign_student(diploy_list)


if __name__ == '__main__':
    print('提示：石墨表格下好再进行下一步哦BBY~')
    os.system('pause')
    wb = openpyxl.load_workbook('账号信息.xlsx')
    crm = wb.worksheets[0]
    crm_user, crm_pwd = str(crm.cell(2,1).value), str(crm.cell(2,2).value)
    download_path = str(crm.cell(2,3).value)
    wb.save('账号信息.xlsx')
    document_name = download_excel(download_path)
    if not document_name:
        print('没有找到 实验班.xlsx，程序退出')
        os.system('pause')
    else:
        # allocated: 实验班的分配结果
        allocated = search_in_excel(document_name, crm_user, crm_pwd)
        print('实验班数据：', allocated)
        deploy(allocated, crm_user, crm_pwd)
        os.system('pause')

   

