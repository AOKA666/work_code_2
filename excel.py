import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import random

from info import get_teacher_dict


class Excel:
    """更新石墨中的表格"""
    def __init__(self, document_name) -> None:
        self.teacher_dict = get_teacher_dict()
        self.allocated = {} # [{'老师id':'学校id'},{'老师id':'学校id'}]
        self.document_name = document_name
        self.wb = openpyxl.load_workbook(f"{self.document_name}")
        self.ws = self.wb.worksheets[0]
        self.nrows = self.ws.max_row
        self.font = Font(
            name="微软雅黑",   # 字体
            size=10,         # 字体大小
        )
        self.alignment = Alignment(horizontal='center',)

    def get_random_teacher(self):
        return random.choice(list(self.teacher_dict.keys()))

    def search(self, data_list):
        """查询学校，分配认证老师"""
        u_list = []
        # 石墨中所有的实验班列表
        for row in range(2, self.nrows+1):
            name = self.ws.cell(row, 1).value
            if not name:
                continue
            u_list.append(name)
        target_row = len(u_list)+2 # 若新增数据，行数
        # 匹配当前拿到的列表
        shiyanban = 0
        for data in data_list:
            school_name = data['school']
            order_id = data['id']
            if school_name == '':
                continue
            shiyanban += 1
            if school_name in u_list:
                index = u_list.index(school_name) + 2
                teacher = self.ws.cell(index, 3).value
                if teacher:
                    # 本来有认证老师
                    teacher_id = self.teacher_dict.get(teacher)
                else:
                    # 没有认证老师
                    teacher = self.get_random_teacher()
                    self.ws.cell(index, 3).value = teacher
                    teacher_id = self.teacher_dict.get(teacher)
                print(teacher, data['school'], f"第{index}行 +1个学员")
                current = self.ws.cell(index, 4).value
                if current:
                    if type(current) == float or type(current) == int:
                        self.ws.cell(index, 4).value = str(int(self.ws.cell(index, 4).value) + 1)
                    else:
                        self.ws.cell(index, 4).value = current + '+1'
                else:
                    self.ws.cell(index, 4).value = '=1'
                    self.ws.cell(target_row, i).font = self.font
                    self.ws.cell(target_row, i).alignment = self.alignment
            else:
                # 新增一行数据
                allocate_teacher = self.get_random_teacher()
                teacher_id = self.teacher_dict.get(allocate_teacher)            
                self.ws.cell(target_row, 1).value = school_name
                self.ws.cell(target_row, 1).font = self.font
                self.ws.cell(target_row, 1).alignment = Alignment(horizontal='left',)
                self.ws.cell(target_row, 2).value = '高校-实验班'
                self.ws.cell(target_row, 3).value = allocate_teacher
                self.ws.cell(target_row, 4).value = 1
                for i in range(2,5):
                    self.ws.cell(target_row, i).font = self.font
                    self.ws.cell(target_row, i).alignment = self.alignment
                print(allocate_teacher, data['school'], f"第{target_row}行 +1个学员")
                # 更新学校列表
                u_list.append(school_name)
                target_row += 1
            if not self.allocated.get(teacher_id):
                self.allocated[teacher_id] = [order_id]
            else:
                self.allocated[teacher_id].append(order_id)
        print("共有{}个实验班".format(shiyanban))
        print("表格更新完成")
        self.wb.save('BBY.xlsx')
        return self.allocated

